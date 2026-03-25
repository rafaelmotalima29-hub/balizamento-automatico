"""
Balizamento: seeded XLSX export (one tab per competition group).

Sheet layout (one row per lane):
  Série | Raia | Nome | Matrícula | Ano Escolar | Sala | Minutos | Segundos | Centésimos
"""

import io
from collections import defaultdict
from flask import Blueprint, render_template, send_file, request
from flask_login import login_required
from models import Student, Event, Group
from services.seeding import build_series, build_relay_series, COMPETITION_GROUPS, YEAR_TO_GROUP

balizamento_bp = Blueprint("balizamento", __name__)

BASE_COLS = ["Série", "Raia", "Nome", "Matrícula", "Ano Escolar", "Sala"]
TIME_COLS = ["Minutos", "Segundos", "Centésimos"]
HEADER = BASE_COLS + TIME_COLS

# ── Colour palette ────────────────────────────────────────────────────

TAB_COLOURS = {
    "6º e 7º Ano":  "163340",
    "8º e 9º Ano":  "163322",
    "Ensino Médio": "2D1A40",
}
SERIE_HEADER_COLOURS = {
    "6º e 7º Ano":  "0D2530",
    "8º e 9º Ano":  "0D2518",
    "Ensino Médio": "1E1030",
}
HEADER_BG      = "1C2230"
HEADER_FG      = "FFFFFF"
ACCENT_FG      = "00C9B1"
ODD_ROW_BG     = "1A2030"
EVEN_ROW_BG    = "141824"
SERIE_SEP_BG   = "0A1018"


# ── Route: preview page ───────────────────────────────────────────────

@balizamento_bp.route("/balizamento")
@login_required
def balizamento():
    selected_event_id = request.args.get("event_id", type=int)
    selected_base_id  = request.args.get("base_id",  type=int)

    all_events = Event.query.order_by(Event.name).all()
    all_bases  = Group.query.order_by(Group.name).all()

    has_selection = bool(selected_event_id or selected_base_id)
    preview     = {}
    has_data    = False
    group_stats = {}
    students    = []
    events      = []

    if has_selection:
        # Filter students by base (if chosen)
        sq = Student.query.order_by(Student.school_year, Student.full_name)
        if selected_base_id:
            sq = sq.filter_by(group_id=selected_base_id)
        students = sq.all()

        # Filter events by id (if chosen)
        events = all_events
        if selected_event_id:
            events = [e for e in all_events if e.id == selected_event_id]

        preview  = _build_preview(events, students)
        has_data = bool(students and events)

        for group in COMPETITION_GROUPS:
            years = _years_for_group(group)
            group_stats[group] = {
                "students": len([s for s in students if s.school_year in years]),
                "events":   len([e for e in events
                                 if e.competition_group and
                                 group in [x.strip() for x in e.competition_group.split(",")]]),
            }

    return render_template(
        "balizamento.html",
        all_events=all_events,
        all_bases=all_bases,
        selected_event_id=selected_event_id,
        selected_base_id=selected_base_id,
        has_selection=has_selection,
        students=students,
        events=events,
        preview=preview,
        has_data=has_data,
        competition_groups=COMPETITION_GROUPS,
        group_stats=group_stats,
    )


# ── Route: XLSX export ────────────────────────────────────────────────

@balizamento_bp.route("/balizamento/export/xlsx")
@login_required
def export_xlsx():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    event_id = request.args.get("event_id", type=int)
    base_id  = request.args.get("base_id",  type=int)

    # Filter students
    sq = Student.query.order_by(Student.school_year, Student.full_name)
    if base_id:
        sq = sq.filter_by(group_id=base_id)
    students = sq.all()

    # Filter events
    all_events = Event.query.order_by(Event.name).all()
    if event_id:
        events = [e for e in all_events if e.id == event_id]
    else:
        events = all_events

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    grouped = defaultdict(list)
    for ev in events:
        if ev.competition_group:
            for g in ev.competition_group.split(","):
                grouped[g.strip()].append(ev)
        else:
            grouped["Sem Grupo"].append(ev)

    sheet_order = [g for g in COMPETITION_GROUPS if g in grouped]
    if "Sem Grupo" in grouped:
        sheet_order.append("Sem Grupo")

    for group_name in sheet_order:
        group_events = grouped[group_name]

        group_years    = _years_for_group(group_name)
        group_students = [s for s in students if s.school_year in group_years] if group_years else students

        ws = wb.create_sheet(title=_safe_sheet_name(group_name))
        tab_colour = TAB_COLOURS.get(group_name, "1C2230")
        ws.sheet_properties.tabColor = tab_colour

        # ── Styles ──────────────────────────────────────────────────
        header_fill    = PatternFill("solid", fgColor=HEADER_BG)
        accent_fill    = PatternFill("solid", fgColor=tab_colour)
        odd_fill       = PatternFill("solid", fgColor=ODD_ROW_BG)
        even_fill      = PatternFill("solid", fgColor=EVEN_ROW_BG)
        serie_hdr_bg   = SERIE_HEADER_COLOURS.get(group_name, "0D1828")
        serie_hdr_fill = PatternFill("solid", fgColor=serie_hdr_bg)
        sep_fill       = PatternFill("solid", fgColor=SERIE_SEP_BG)

        h_font_base = Font(bold=True, color=ACCENT_FG, size=10)
        h_font_time = Font(bold=True, color=HEADER_FG, size=10)
        data_font   = Font(color=HEADER_FG, size=10)
        serie_font  = Font(bold=True, color=ACCENT_FG, size=9, italic=True)
        center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align  = Alignment(horizontal="left",   vertical="center")

        thin = Border(
            bottom=Side(border_style="thin", color="2D3748"),
            right=Side(border_style="thin",  color="2D3748"),
        )
        thick_top = Border(
            top=Side(border_style="medium", color="00C9B1"),
            bottom=Side(border_style="thin", color="2D3748"),
            right=Side(border_style="thin",  color="2D3748"),
        )

        current_row = 1

        for ev in group_events:
            # ── Event title row ──────────────────────────────────────
            title_cell = ws.cell(row=current_row, column=1,
                                 value=f"🏁  {ev.name}   ({ev.num_series} série(s) × {ev.athletes_per_series} raias)")
            title_cell.font      = Font(bold=True, color=ACCENT_FG, size=12)
            title_cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
            title_cell.alignment = left_align
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=len(HEADER))
            ws.row_dimensions[current_row].height = 30
            current_row += 1

            # ── Column header row ────────────────────────────────────
            for col_idx, col_name in enumerate(HEADER, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=col_name)
                cell.fill      = accent_fill if col_idx <= len(BASE_COLS) else header_fill
                cell.font      = h_font_base if col_idx <= len(BASE_COLS) else h_font_time
                cell.alignment = center
                cell.border    = thin
            ws.row_dimensions[current_row].height = 36
            ws.freeze_panes = ws.cell(row=current_row + 1, column=3)
            current_row += 1

            # ── Seeded data rows ─────────────────────────────────────
            ev_students = [s for s in group_students
                           if ev.group_id is None or s.group_id == ev.group_id]
            ev_students = _filter_by_event_gender(ev_students, ev)

            if ev.is_relay:
                all_series = build_relay_series(ev, ev_students, event_group=group_name)
            else:
                all_series = build_series(ev, ev_students, event_group=group_name)

            for series_idx, series in enumerate(all_series, start=1):
                # ── Serie header row ─────────────────────────────────
                serie_label = f"  Série {series_idx} de {len(all_series)}"
                serie_cell  = ws.cell(row=current_row, column=1, value=serie_label)
                serie_cell.font      = serie_font
                serie_cell.fill      = serie_hdr_fill
                serie_cell.alignment = left_align
                ws.merge_cells(start_row=current_row, start_column=1,
                               end_row=current_row, end_column=len(HEADER))
                ws.row_dimensions[current_row].height = 16
                current_row += 1

                if ev.is_relay:
                    # ── Relay: each lane is a team ──────────────────
                    relay_size = ev.relay_size or 4
                    for lane_idx, team in enumerate(series, start=1):
                        # Team header row
                        team_label = f"  Raia {lane_idx} — {team['year']}" if team else f"  Raia {lane_idx} — (vaga livre)"
                        team_cell = ws.cell(row=current_row, column=1, value=team_label)
                        team_cell.font = Font(bold=True, color=ACCENT_FG, size=10)
                        team_cell.fill = serie_hdr_fill
                        team_cell.alignment = left_align
                        ws.merge_cells(start_row=current_row, start_column=1,
                                       end_row=current_row, end_column=len(HEADER))
                        ws.row_dimensions[current_row].height = 20
                        current_row += 1

                        # Member rows
                        members = team["students"] if team else []
                        for m_idx in range(relay_size):
                            student = members[m_idx] if m_idx < len(members) else None
                            row_fill = odd_fill if (current_row % 2 == 0) else even_fill
                            row_data = [
                                series_idx,
                                lane_idx,
                                student.full_name    if student else "",
                                student.registration if student else "",
                                student.school_year  if student else "",
                                student.classroom or "" if student else "",
                            ]
                            # Time columns only on last member row (team time)
                            if m_idx == relay_size - 1:
                                row_data += ["", "", ""]  # Minutos, Segundos, Centésimos
                            else:
                                row_data += ["", "", ""]
                            use_thick = m_idx == 0
                            for col_idx, value in enumerate(row_data, start=1):
                                cell = ws.cell(row=current_row, column=col_idx, value=value)
                                cell.fill      = row_fill
                                cell.font      = data_font
                                cell.border    = thick_top if use_thick else thin
                                cell.alignment = center if col_idx <= 2 else left_align
                            ws.row_dimensions[current_row].height = 22
                            current_row += 1
                else:
                    # ── Individual: one student per lane ─────────────
                    for lane_idx, student in enumerate(series, start=1):
                        row_fill = odd_fill if (current_row % 2 == 0) else even_fill
                        row_data = [
                            series_idx,
                            lane_idx,
                            student.full_name    if student else "",
                            student.registration if student else "",
                            student.school_year  if student else "",
                            student.classroom or "" if student else "",
                            "",  # Minutos
                            "",  # Segundos
                            "",  # Centésimos
                        ]
                        use_thick = lane_idx == 1
                        for col_idx, value in enumerate(row_data, start=1):
                            cell = ws.cell(row=current_row, column=col_idx, value=value)
                            cell.fill      = row_fill
                            cell.font      = data_font
                            cell.border    = thick_top if use_thick else thin
                            cell.alignment = center if col_idx <= 2 else left_align
                        ws.row_dimensions[current_row].height = 22
                        current_row += 1

                # ── Spacer between series ────────────────────────────
                if series_idx < len(all_series):
                    for col_idx in range(1, len(HEADER) + 1):
                        sep_cell      = ws.cell(row=current_row, column=col_idx)
                        sep_cell.fill = sep_fill
                    ws.row_dimensions[current_row].height = 5
                    current_row += 1

            # ── Blank spacer between events ──────────────────────────
            current_row += 2

        # ── Column widths ────────────────────────────────────────────
        col_widths = [8, 8, 28, 14, 18, 10, 12, 12, 14]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Build descriptive filename
    parts = []
    if event_id:
        ev_name = next((e.name for e in all_events if e.id == event_id), str(event_id))
        parts.append(ev_name.replace(" ", "_"))
    if base_id:
        base_name = next((b.name for b in Group.query.all() if b.id == base_id), str(base_id))
        parts.append(base_name.replace(" ", "_").replace("–", "-"))
    filename = f"balizamento_{'_'.join(parts)}.xlsx" if parts else "balizamento_completo.xlsx"

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ── Helpers ───────────────────────────────────────────────────────────

def _filter_by_event_gender(students, event):
    """Filter students by event gender (M/F/MISTO/None)."""
    if not event.gender or event.gender == "MISTO":
        return students
    return [s for s in students if s.gender == event.gender]


def _build_preview(events, students):
    """Return {group: [(event, series_list, is_relay)]} for template preview."""
    grouped = defaultdict(list)
    for ev in events:
        if ev.competition_group:
            for g in ev.competition_group.split(","):
                key = g.strip()
                group_years    = _years_for_group(key)
                group_students = [s for s in students if s.school_year in group_years] if group_years else students
                ev_students    = [s for s in group_students if ev.group_id is None or s.group_id == ev.group_id]
                ev_students    = _filter_by_event_gender(ev_students, ev)
                if ev.is_relay:
                    series_list = build_relay_series(ev, ev_students, event_group=key)
                else:
                    series_list = build_series(ev, ev_students, event_group=key)
                grouped[key].append((ev, series_list, ev.is_relay))
        else:
            key = "Sem Grupo"
            ev_students = [s for s in students if ev.group_id is None or s.group_id == ev.group_id]
            ev_students = _filter_by_event_gender(ev_students, ev)
            if ev.is_relay:
                series_list = build_relay_series(ev, ev_students, event_group=key)
            else:
                series_list = build_series(ev, ev_students, event_group=key)
            grouped[key].append((ev, series_list, ev.is_relay))
    return dict(grouped)


def _years_for_group(group: str) -> list[str]:
    return [y for y, g in YEAR_TO_GROUP.items() if g == group]


def _safe_sheet_name(name: str) -> str:
    bad   = r'\/*?:[]'
    clean = "".join(c for c in name if c not in bad)
    return clean[:31]
