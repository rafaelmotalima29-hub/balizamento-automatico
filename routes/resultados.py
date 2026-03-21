"""
Resultados: detailed competition results dashboard.

Groups results by competition_group and provides:
  - Podium (top 3 overall per group)
  - Full group ranking
  - Per-school-year sub-rankings
  - XLSX export with multiple sheets
"""

import io
from collections import defaultdict
from flask import Blueprint, render_template, send_file

from models import Student, Event, Result, ScoreConfig
from extensions import db
from sqlalchemy import func

resultados_bp = Blueprint("resultados", __name__)


@resultados_bp.route("/resultados")
def resultados():
    has_results = Result.query.count() > 0
    if not has_results:
        return render_template("resultados.html", groups=[], has_results=False)

    # Load all results with students/events
    results = (
        Result.query
        .filter(Result.is_dq == False, Result.total_time.isnot(None))
        .join(Student)
        .join(Event)
        .order_by(Result.total_time.asc())
        .all()
    )

    # Build per-group data
    # {competition_group: [result, ...]}
    from services.seeding import YEAR_TO_GROUP
    by_group: dict[str, list] = defaultdict(list)
    for r in results:
        if r.event.competition_group:
            event_groups = [g.strip() for g in r.event.competition_group.split(",")]
            stu_group = YEAR_TO_GROUP.get(r.student.school_year)
            if stu_group in event_groups:
                by_group[stu_group].append(r)
            else:
                by_group[event_groups[0]].append(r)
        else:
            by_group["Sem Grupo"].append(r)

    # Order groups
    from services.seeding import COMPETITION_GROUPS
    group_order = [g for g in COMPETITION_GROUPS if g in by_group]
    if "Sem Grupo" in by_group and "Sem Grupo" not in group_order:
        group_order.append("Sem Grupo")

    groups_data = []
    for group_name in group_order:
        group_results = by_group[group_name]
        # Sort by total_time ascending
        group_results.sort(key=lambda r: r.total_time)

        # Podium: top 3 unique students (by best time)
        seen_students = set()
        podium = []
        for r in group_results:
            if r.student_id not in seen_students:
                seen_students.add(r.student_id)
                podium.append(r)
            if len(podium) == 3:
                break

        # Full ranking: deduplicate students (best time wins)
        seen2 = set()
        full_ranking = []
        rank_counter = 1
        prev_time = None
        for r in group_results:
            if r.student_id in seen2:
                continue
            seen2.add(r.student_id)
            if r.total_time != prev_time:
                current_rank = rank_counter
                prev_time = r.total_time
            full_ranking.append({"rank": current_rank, "result": r})
            rank_counter += 1

        # Per-year rankings
        year_rankings: dict[str, list] = defaultdict(list)
        for entry in full_ranking:
            yr = entry["result"].student.school_year
            year_rankings[yr].append(entry)

        # Re-rank within each year (with tie handling)
        for yr in year_rankings:
            entries = year_rankings[yr]
            entries.sort(key=lambda e: e["result"].total_time)
            prev_time = None
            yr_rank_counter = 1
            current_yr_rank = 1
            for i, entry in enumerate(entries):
                t = entry["result"].total_time
                if t != prev_time:
                    current_yr_rank = yr_rank_counter
                    prev_time = t
                entry["year_rank"] = current_yr_rank
                yr_rank_counter += 1

        groups_data.append({
            "name": group_name,
            "podium": podium,
            "full_ranking": full_ranking,
            "year_rankings": dict(year_rankings),
            "years": sorted(year_rankings.keys()),
        })

    return render_template(
        "resultados.html",
        groups=groups_data,
        has_results=True,
    )


def _fmt_time(total: float) -> str:
    """Format float seconds as M:SS.cc"""
    m = int(total // 60)
    s = int(total % 60)
    c = round((total - int(total)) * 100)
    return f"{m}:{s:02d}.{c:02d}"


# ── XLSX Export ──────────────────────────────────────────────────────

@resultados_bp.route("/resultados/export/xlsx")
def export_xlsx():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from services.seeding import COMPETITION_GROUPS, YEAR_TO_GROUP

    # ── Colour palette ───────────────────────────────────────────────
    BG_DARK      = "0E1117"
    BG_CARD      = "161B22"
    BG_HEADER    = "1C2230"
    BG_ODD       = "1A2030"
    BG_EVEN      = "141824"
    FG_WHITE     = "FFFFFF"
    FG_ACCENT    = "00C9B1"
    FG_GOLD      = "FFD700"
    FG_SILVER    = "C0C0C0"
    FG_BRONZE    = "CD7F32"
    BORDER_CLR   = "2D3748"

    TAB_COLOURS = {
        "Resultado Geral": "163340",
        "6º e 7º Ano":    "163340",
        "8º e 9º Ano":    "163322",
        "Ensino Médio":    "2D1A40",
        "Pontuação":       "1A3320",
    }
    YEAR_TAB_COLOURS = {
        "6º Ano": "0D2530", "7º Ano": "0D2838",
        "8º Ano": "0D2518", "9º Ano": "0D2820",
        "1º Ano Médio": "1E1030", "2º Ano Médio": "241535", "3º Ano Médio": "2A1A40",
    }

    # ── Style helpers ────────────────────────────────────────────────
    thin_border = Border(
        bottom=Side(border_style="thin", color=BORDER_CLR),
        right=Side(border_style="thin", color=BORDER_CLR),
    )
    header_fill   = PatternFill("solid", fgColor=BG_HEADER)
    odd_fill      = PatternFill("solid", fgColor=BG_ODD)
    even_fill     = PatternFill("solid", fgColor=BG_EVEN)
    accent_font   = Font(bold=True, color=FG_ACCENT, size=10)
    white_font    = Font(color=FG_WHITE, size=10)
    bold_white    = Font(bold=True, color=FG_WHITE, size=10)
    title_font    = Font(bold=True, color=FG_ACCENT, size=13)
    subtitle_font = Font(bold=True, color=FG_ACCENT, size=11)
    center        = Alignment(horizontal="center", vertical="center")
    left_align    = Alignment(horizontal="left", vertical="center")

    medal_fonts = {
        1: Font(bold=True, color=FG_GOLD, size=10),
        2: Font(bold=True, color=FG_SILVER, size=10),
        3: Font(bold=True, color=FG_BRONZE, size=10),
    }

    def _style_header_row(ws, row, cols, _ncols=None):
        for ci, col_name in enumerate(cols, 1):
            cell = ws.cell(row=row, column=ci, value=col_name)
            cell.fill = header_fill
            cell.font = accent_font
            cell.alignment = center
            cell.border = thin_border
        ws.row_dimensions[row].height = 28

    def _write_data_row(ws, row, values, rank=None):
        fill = odd_fill if row % 2 == 0 else even_fill
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = center if ci in (1,) else left_align
            if ci == 1 and rank and rank <= 3:
                cell.font = medal_fonts[rank]
            elif ci == 1:
                cell.font = accent_font
            elif ci == len(values):  # time column
                cell.font = Font(bold=True, color=FG_ACCENT, size=10)
                cell.alignment = center
            else:
                cell.font = white_font
        ws.row_dimensions[row].height = 22

    def _write_title(ws, row, text, ncols):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = title_font
        cell.fill = PatternFill("solid", fgColor=BG_DARK)
        cell.alignment = left_align
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        ws.row_dimensions[row].height = 32

    def _write_subtitle(ws, row, text, ncols):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = subtitle_font
        cell.fill = PatternFill("solid", fgColor=BG_CARD)
        cell.alignment = left_align
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        ws.row_dimensions[row].height = 26

    def _set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _safe_name(name):
        bad = r'\/*?:[]'
        clean = "".join(c for c in name if c not in bad)
        return clean[:31]

    def _rank_label(rank):
        if rank == 1: return "🥇 1º"
        if rank == 2: return "🥈 2º"
        if rank == 3: return "🥉 3º"
        return f"{rank}º"

    # ── Load data ────────────────────────────────────────────────────
    results = (
        Result.query
        .filter(Result.is_dq == False, Result.total_time.isnot(None))
        .join(Student)
        .join(Event)
        .order_by(Result.total_time.asc())
        .all()
    )

    if not results:
        # Return empty workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sem Resultados"
        ws.cell(row=1, column=1, value="Nenhum resultado encontrado.")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="resultados_competicao.xlsx")

    # ── Build ranking data (same logic as the view) ──────────────────
    by_group = defaultdict(list)
    for r in results:
        if r.event.competition_group:
            event_groups = [g.strip() for g in r.event.competition_group.split(",")]
            stu_group = YEAR_TO_GROUP.get(r.student.school_year)
            if stu_group in event_groups:
                by_group[stu_group].append(r)
            else:
                by_group[event_groups[0]].append(r)
        else:
            by_group["Sem Grupo"].append(r)

    group_order = [g for g in COMPETITION_GROUPS if g in by_group]
    if "Sem Grupo" in by_group and "Sem Grupo" not in group_order:
        group_order.append("Sem Grupo")

    groups_data = []
    all_ranking = []  # flat list for the general sheet

    for group_name in group_order:
        group_results = sorted(by_group[group_name], key=lambda r: r.total_time)

        seen = set()
        full_ranking = []
        rank_counter = 1
        prev_time = None
        for r in group_results:
            if r.student_id in seen:
                continue
            seen.add(r.student_id)
            if r.total_time != prev_time:
                current_rank = rank_counter
                prev_time = r.total_time
            entry = {"rank": current_rank, "result": r, "group": group_name}
            full_ranking.append(entry)
            all_ranking.append(entry)
            rank_counter += 1

        year_rankings = defaultdict(list)
        for entry in full_ranking:
            yr = entry["result"].student.school_year
            year_rankings[yr].append(entry)

        for yr in year_rankings:
            entries = year_rankings[yr]
            entries.sort(key=lambda e: e["result"].total_time)
            prev_t = None
            yr_rank = 1
            for i, e in enumerate(entries):
                t = e["result"].total_time
                if t != prev_t:
                    current_yr_rank = yr_rank
                    prev_t = t
                e["year_rank"] = current_yr_rank
                yr_rank += 1

        groups_data.append({
            "name": group_name,
            "full_ranking": full_ranking,
            "year_rankings": dict(year_rankings),
            "years": sorted(year_rankings.keys()),
        })

    # ── Team scoring ─────────────────────────────────────────────────
    team_scores = (
        db.session.query(
            Student.school_year,
            func.sum(Result.points).label("total_points"),
            func.count(Result.id).label("total_results"),
        )
        .join(Result, Result.student_id == Student.id)
        .group_by(Student.school_year)
        .order_by(func.sum(Result.points).desc())
        .all()
    )

    score_config = {sc.placement: sc.points for sc in ScoreConfig.query.all()}

    # ══════════════════════════════════════════════════════════════════
    # BUILD THE WORKBOOK
    # ══════════════════════════════════════════════════════════════════
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    RANK_COLS = ["Pos.", "Atleta", "Matrícula", "Ano / Equipe", "Sala", "Prova", "Tempo"]
    RANK_WIDTHS = [8, 30, 14, 18, 10, 22, 12]
    YEAR_COLS = ["Pos.", "Atleta", "Matrícula", "Sala", "Prova", "Tempo"]
    YEAR_WIDTHS = [8, 30, 14, 10, 22, 12]

    # ── Sheet 1: Resultado Geral ─────────────────────────────────────
    ws = wb.create_sheet(title="Resultado Geral")
    ws.sheet_properties.tabColor = TAB_COLOURS.get("Resultado Geral", "163340")
    row = 1

    _write_title(ws, row, "🏆  Resultado Geral da Competição", len(RANK_COLS))
    row += 2

    # Re-rank the combined all_ranking by total_time
    all_ranking.sort(key=lambda e: e["result"].total_time)
    prev_time = None
    gen_rank = 1
    for i, entry in enumerate(all_ranking):
        t = entry["result"].total_time
        if t != prev_time:
            current_gen_rank = gen_rank
            prev_time = t
        entry["gen_rank"] = current_gen_rank
        gen_rank += 1

    _style_header_row(ws, row, RANK_COLS, len(RANK_COLS))
    row += 1

    for entry in all_ranking:
        r = entry["result"]
        _write_data_row(ws, row, [
            _rank_label(entry["gen_rank"]),
            r.student.full_name,
            r.student.registration,
            r.student.school_year,
            r.student.classroom or "—",
            r.event.name,
            _fmt_time(r.total_time),
        ], rank=entry["gen_rank"])
        row += 1

    _set_col_widths(ws, RANK_WIDTHS)

    # ── Sheets 2-N: One per competition group ────────────────────────
    for gd in groups_data:
        ws = wb.create_sheet(title=_safe_name(gd["name"]))
        ws.sheet_properties.tabColor = TAB_COLOURS.get(gd["name"], "163340")
        row = 1

        _write_title(ws, row, f"🏊  {gd['name']} — Classificação Geral", len(RANK_COLS))
        row += 2

        _style_header_row(ws, row, RANK_COLS, len(RANK_COLS))
        row += 1

        for entry in gd["full_ranking"]:
            r = entry["result"]
            _write_data_row(ws, row, [
                _rank_label(entry["rank"]),
                r.student.full_name,
                r.student.registration,
                r.student.school_year,
                r.student.classroom or "—",
                r.event.name,
                _fmt_time(r.total_time),
            ], rank=entry["rank"])
            row += 1

        _set_col_widths(ws, RANK_WIDTHS)

    # ── Sheets: One per school year ──────────────────────────────────
    for gd in groups_data:
        for year in gd["years"]:
            entries = gd["year_rankings"][year]
            sheet_name = _safe_name(f"{year}")
            # Avoid duplicate sheet names
            existing = [ws.title for ws in wb.worksheets]
            if sheet_name in existing:
                sheet_name = _safe_name(f"{year} ({gd['name'][:10]})")

            ws = wb.create_sheet(title=sheet_name)
            ws.sheet_properties.tabColor = YEAR_TAB_COLOURS.get(year, "1C2230")
            row = 1

            _write_title(ws, row, f"🎓  {year} — {gd['name']}", len(YEAR_COLS))
            row += 2

            _style_header_row(ws, row, YEAR_COLS, len(YEAR_COLS))
            row += 1

            for entry in entries:
                r = entry["result"]
                _write_data_row(ws, row, [
                    _rank_label(entry["year_rank"]),
                    r.student.full_name,
                    r.student.registration,
                    r.student.classroom or "—",
                    r.event.name,
                    _fmt_time(r.total_time),
                ], rank=entry["year_rank"])
                row += 1

            _set_col_widths(ws, YEAR_WIDTHS)

    # ── Sheet: Pontuação (team scoring) ──────────────────────────────
    ws = wb.create_sheet(title="Pontuação")
    ws.sheet_properties.tabColor = TAB_COLOURS.get("Pontuação", "1A3320")
    row = 1

    _write_title(ws, row, "🏅  Pontuação por Equipe", 4)
    row += 2

    score_cols = ["Pos.", "Equipe (Ano Escolar)", "Total de Pontos", "Nº Resultados"]
    _style_header_row(ws, row, score_cols, len(score_cols))
    row += 1

    for i, ts in enumerate(team_scores, 1):
        _write_data_row(ws, row, [
            _rank_label(i),
            ts.school_year,
            int(ts.total_points),
            int(ts.total_results),
        ], rank=i)
        row += 1

    # Scoring table reference
    row += 2
    _write_subtitle(ws, row, "📋  Tabela de Pontuação Utilizada", 4)
    row += 1

    ref_cols = ["Colocação", "Pontos", "", ""]
    _style_header_row(ws, row, ref_cols, 4)
    row += 1

    for placement in sorted(score_config.keys()):
        fill = odd_fill if row % 2 == 0 else even_fill
        c1 = ws.cell(row=row, column=1, value=f"{placement}º lugar")
        c1.fill = fill; c1.font = accent_font; c1.alignment = center; c1.border = thin_border
        c2 = ws.cell(row=row, column=2, value=f"{score_config[placement]} pts")
        c2.fill = fill; c2.font = bold_white; c2.alignment = center; c2.border = thin_border
        for ci in (3, 4):
            c = ws.cell(row=row, column=ci)
            c.fill = fill; c.border = thin_border
        ws.row_dimensions[row].height = 22
        row += 1

    _set_col_widths(ws, [10, 24, 16, 16])

    # ── Save and return ──────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="resultados_competicao.xlsx",
    )
